import openpyxl
import pytest

from levi.evidence.models import EvidenceType
from levi.evidence.parsers.base import ParserValidationError
from levi.evidence.parsers.excel_parser import ExcelEvidenceParser


def _workbook(tmp_path, setup):
    path = tmp_path / "book.xlsx"
    workbook = openpyxl.Workbook()
    setup(workbook)
    workbook.save(path)
    return path


def _parse(path):
    return ExcelEvidenceParser().parse(file_path=path, user_id="u1", source_name="Upload")


def test_multiple_sheets_parsed(tmp_path):
    def setup(wb):
        wb.active.append(["symbol", "quantity", "market_value"])
        wb.active.append(["SPY", 1, 100])
        second = wb.create_sheet("Second")
        second.append(["name", "value"])
        second.append(["x", 1])
    parsed = _parse(_workbook(tmp_path, setup))
    assert len(parsed.structured_data["sheets"]) == 2


def test_empty_sheet_skipped(tmp_path):
    def setup(wb):
        wb.active.append(["symbol", "quantity", "market_value"])
        wb.active.append(["SPY", 1, 100])
        wb.create_sheet("Empty")
    parsed = _parse(_workbook(tmp_path, setup))
    assert any("Empty sheet" in warning for warning in parsed.warnings)


def test_hidden_sheet_warning(tmp_path):
    def setup(wb):
        wb.active.append(["symbol", "quantity", "market_value"])
        wb.active.append(["SPY", 1, 100])
        hidden = wb.create_sheet("Hidden")
        hidden.append(["secret"])
        hidden.sheet_state = "hidden"
    parsed = _parse(_workbook(tmp_path, setup))
    assert any("Hidden sheet" in warning for warning in parsed.warnings)


def test_formula_not_executed(tmp_path):
    def setup(wb):
        wb.active.append(["name", "value"])
        wb.active.append(["total", "=1+1"])
    parsed = _parse(_workbook(tmp_path, setup))
    assert parsed.structured_data["sheets"][0]["rows"][0]["value"] == "=1+1"


def test_options_chain_sheet_detected(tmp_path):
    def setup(wb):
        wb.active.append(["symbol", "expiration", "strike", "bid", "ask"])
        wb.active.append(["SPY", "2026-08-21", 600, 1, 1.1])
    parsed = _parse(_workbook(tmp_path, setup))
    assert parsed.evidence_type is EvidenceType.OPTIONS_CHAIN


def test_portfolio_sheet_detected(tmp_path):
    def setup(wb):
        wb.active.append(["symbol", "quantity", "market_value", "cost_basis"])
        wb.active.append(["AAPL", 2, 400, 350])
    parsed = _parse(_workbook(tmp_path, setup))
    assert parsed.evidence_type is EvidenceType.PORTFOLIO_EXPORT
    assert parsed.ticker_symbols == ("AAPL",)


def test_unsupported_workbook_rejected(tmp_path):
    path = tmp_path / "bad.xls"
    path.write_bytes(b"not a workbook")
    with pytest.raises(ParserValidationError, match="invalid or unsupported XLS"):
        _parse(path)


def test_xlsm_not_supported():
    assert not ExcelEvidenceParser().supports(
        filename="macro.xlsm", mime_type="application/vnd.ms-excel"
    )


def test_excel_timestamps_extracted_conservatively(tmp_path):
    def setup(wb):
        wb.active.append(["symbol", "entry_time", "entry_price", "exit_price"])
        wb.active.append(["SPY", "2026-07-21T10:00:00", 10, 11])
    parsed = _parse(_workbook(tmp_path, setup))
    assert parsed.structured_data["parsed_timestamps"] == ["2026-07-21T10:00:00"]
